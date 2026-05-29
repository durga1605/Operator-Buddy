"""Views for rendering main menu and navigation pages in the Mobility module."""

from datetime import datetime
from django.shortcuts import render
from django.http import HttpResponse
import openpyxl
from ..components.db_connection_string import get_db_connection


def home_page(request):
    """Render the main mobility menu page."""
    return render(request, "home/home.html")


def inspection_menu(request):
    """Render the inspection menu page."""
    return render(request, "inspection/inspection_menu.html")


def view_logs(request):
    """
    Renders the logs page with optional date filtering.
    """
    plant_code = ""
    if not plant_code:
        plant_code = request.session.get("plant_code")

    if not plant_code:
        # Cannot log without knowing which DB to use
        raise ValueError("Plant code is required for logging traceability")
    db = get_db_connection(plant_code)
    logs_collection = db["traceability_log"]

    start_date = None
    end_date = None
    filter_query = {}

    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    if start_date and end_date:
        try:
            start_dt = datetime.strptime(start_date, "%Y-%m-%d")
            end_dt = datetime.strptime(end_date, "%Y-%m-%d")
            # Include whole day till 23:59:59
            end_dt = end_dt.replace(hour=23, minute=59, second=59)
            filter_query["timestamp"] = {"$gte": start_dt, "$lte": end_dt}
        except ValueError:
            # Invalid date format - ignore filter or handle error
            pass

    logs = list(logs_collection.find(filter_query).sort("timestamp", -1))

    return render(
        request,
        "auth/logs.html",
        logs=logs,
        start_date=start_date,
        end_date=end_date,
    )


def download_logs_excel(request):
    """
    Downloads the logs as an Excel file with optional date filtering.
    """
    plant_code = ""
    if not plant_code:
        plant_code = request.session.get("plant_code")

    if not plant_code:
        # Cannot log without knowing which DB to use
        raise ValueError("Plant code is required for logging traceability")
    db = get_db_connection(plant_code)
    logs_collection = db["traceability_log"]

    start_date = None
    end_date = None
    filter_query = {}

    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    if start_date and end_date:
        try:
            start_dt = datetime.strptime(start_date, "%Y-%m-%d")
            end_dt = datetime.strptime(end_date, "%Y-%m-%d")
            # Include whole day till 23:59:59
            end_dt = end_dt.replace(hour=23, minute=59, second=59)
            filter_query["timestamp"] = {"$gte": start_dt, "$lte": end_dt}
        except ValueError:
            # Invalid date format - ignore filter or handle error
            pass

    logs = list(logs_collection.find(filter_query).sort("timestamp", -1))

    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Logs"

    # Add headers
    headers = ["Timestamp", "User", "Name", "IP", "Level", "Message", "Path"]
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)

    # Add data
    for row_num, log in enumerate(logs, 2):
        ws.cell(row=row_num, column=1, value=log.get("timestamp"))
        ws.cell(row=row_num, column=2, value=log.get("user"))
        ws.cell(row=row_num, column=3, value=log.get("name"))
        ws.cell(row=row_num, column=4, value=log.get("ip"))
        ws.cell(row=row_num, column=5, value=log.get("level"))
        ws.cell(row=row_num, column=6, value=log.get("message"))
        ws.cell(row=row_num, column=7, value=log.get("path"))

    # Create response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="logs.xlsx"'

    # Save workbook to response
    wb.save(response)

    return response
